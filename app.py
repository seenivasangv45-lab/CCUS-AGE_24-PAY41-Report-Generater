"""
AGE_24 / PAY_41 Web Processor
==============================
Run this once on the office PC:
    pip install flask pandas openpyxl
    python app.py

Then anyone on the same WiFi can open:
    http://<this-pc-ip>:5000
"""

import sys
import datetime
import io
import os
import re
import tempfile
import threading
import time
import hashlib
import traceback
import warnings
import copy as _copy

import numpy as np
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import LineChart, Reference, Series
from pathlib import Path
from flask import Flask, request, send_file, render_template, jsonify

warnings.filterwarnings("ignore")

app = Flask(__name__)
app.config["MAX_CONTENT_LENGTH"] = 200 * 1024 * 1024   # 200 MB upload limit

# ── Temp-file output store (avoids holding large blobs in RAM) ─────────────────
_OUTPUT_LOCK  = threading.Lock()
_OUTPUT_STORE = {}   # token -> (tmp_file_path, original_filename)

def _save_output(out_bytes: bytes, orig_filename: str) -> str:
    token = hashlib.md5(f"{time.time()}{orig_filename}".encode()).hexdigest()[:12]
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    tmp.write(out_bytes)
    tmp.close()
    with _OUTPUT_LOCK:
        _OUTPUT_STORE[token] = (tmp.name, orig_filename)
    return token

def _get_output(token):
    with _OUTPUT_LOCK:
        return _OUTPUT_STORE.get(token)

def _cleanup_output(token):
    with _OUTPUT_LOCK:
        entry = _OUTPUT_STORE.pop(token, None)
    if entry:
        try:
            os.unlink(entry[0])
        except OSError:
            pass

# ── Constants ─────────────────────────────────────────────────────────────────
EXCLUDE_ALL   = {"11-DMERC", "1-Self Pay", "20-Work Comp", "23-Auto", "9-EPS", "30-WC Non-Responsive"}
WC_MVA_CATS   = {"20-Work Comp", "23-Auto", "30-WC Non-Responsive"}
AMOUNT_COLS   = ["CurrentAmt", "Over30Amt", "Over60Amt", "Over90Amt", "Over120Amt", "Over150Amt"]
PCT_ROWS      = {10, 18, 26, 34, 42, 50}
SUMMARY_SHEET = "Summary"
TEMPLATE_DIR  = Path(__file__).parent / "templates_folder"

# ── Helpers ───────────────────────────────────────────────────────────────────
def bucket_sums(df, mask):
    sub = df.loc[mask, AMOUNT_COLS]
    return sub.sum().tolist()

def total_ar(vals):
    return sum(vals)

def pct_over_120(vals):
    t = total_ar(vals)
    return (vals[4] + vals[5]) / t if t else 0.0

def find_date_column(ws, target_date):
    anchor_val = ws.cell(row=1, column=3).value
    if anchor_val is None:
        return None
    if hasattr(anchor_val, "date"):
        anchor_date = anchor_val.date()
    elif isinstance(anchor_val, datetime.datetime):
        anchor_date = anchor_val.date()
    else:
        try:
            anchor_date = datetime.datetime.strptime(str(anchor_val), "%Y-%m-%d").date()
        except ValueError:
            return None
    delta = (target_date - anchor_date).days
    if delta < 0:
        return None
    return 3 + delta

def copy_cell_style(cell, attr):
    return _copy.copy(getattr(cell, attr))

def load_summary_from_template():
    candidates = list(TEMPLATE_DIR.glob("*.xlsx")) if TEMPLATE_DIR.exists() else []
    if not candidates:
        raise FileNotFoundError(
            f"No template .xlsx found in '{TEMPLATE_DIR}'. "
            "Place a file containing the Summary sheet there."
        )
    tmpl_path = candidates[0]
    wb_tmpl = load_workbook(tmpl_path)
    if SUMMARY_SHEET not in wb_tmpl.sheetnames:
        raise ValueError(
            f"Template '{tmpl_path.name}' has no sheet named '{SUMMARY_SHEET}'."
        )
    return wb_tmpl

# ── Report Sheet Writer ───────────────────────────────────────────────────────
def write_report_sheet(wb, overall, res, start_date=None, end_date=None):
    SHEET_NAME = "AR Report"
    if SHEET_NAME in wb.sheetnames:
        del wb[SHEET_NAME]
    ws = wb.create_sheet(SHEET_NAME)

    _TITLE_FILL = PatternFill("solid", fgColor="F4B942")
    _HDR_FILL   = PatternFill("solid", fgColor="1F4E79")
    _ALT_FILL   = PatternFill("solid", fgColor="DEEAF1")
    _WHITE_FILL = PatternFill("solid", fgColor="FFFFFF")
    _TOTAL_FILL = PatternFill("solid", fgColor="BDD7EE")
    _F_TITLE    = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
    _F_HDR      = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
    _F_BOLD     = Font(name="Calibri", bold=True, size=10)
    _F_NORMAL   = Font(name="Calibri", size=10)
    _THIN       = Side(style="thin", color="B0B0B0")
    _BORDER     = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
    _CENTER     = Alignment(horizontal="center", vertical="center", wrap_text=True)
    _LEFT       = Alignment(horizontal="left",   vertical="center")
    _RIGHT      = Alignment(horizontal="right",  vertical="center")
    _USD        = '"$"#,##0.00_);[Red]("$"#,##0.00)'
    _PCT        = '0%'
    NUM_COLS    = 9

    if start_date and end_date:
        try:
            s = datetime.datetime.strptime(start_date, "%Y-%m-%d").strftime("%m/%d/%Y")
            e = datetime.datetime.strptime(end_date,   "%Y-%m-%d").strftime("%m/%d/%Y")
            title_text = f"{s} to {e}"
        except Exception:
            title_text = f"{start_date} to {end_date}"
    elif start_date:
        try:
            title_text = f"From {datetime.datetime.strptime(start_date, '%Y-%m-%d').strftime('%m/%d/%Y')}"
        except Exception:
            title_text = start_date
    elif end_date:
        try:
            title_text = f"Through {datetime.datetime.strptime(end_date, '%Y-%m-%d').strftime('%m/%d/%Y')}"
        except Exception:
            title_text = end_date
    else:
        title_text = "Enter date range here (e.g. 09/01/2025 to 04/13/2026)"

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=NUM_COLS)
    tc = ws.cell(row=1, column=1, value=title_text)
    tc.font = _F_TITLE; tc.fill = _TITLE_FILL; tc.alignment = _CENTER; tc.border = _BORDER

    headers = ["ITEMS","Total AR (Age -24)","Insurance CURRENT","Insurance 31-60",
               "Insurance 61-90","Insurance 91-120","Insurance 121-150","Insurance 151+","% AR > 120"]
    col_widths = [18, 18, 18, 15, 15, 15, 17, 15, 13]
    for ci, (hdr, w) in enumerate(zip(headers, col_widths), start=1):
        c = ws.cell(row=2, column=ci, value=hdr)
        c.font = _F_HDR; c.fill = _HDR_FILL; c.alignment = _CENTER; c.border = _BORDER
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 30

    rows_data = [
        ("Overall Aging", overall,          True),
        ("Insurance AR",  res["Insurance"], False),
        ("WC & MVA",      res["WC"],        False),
        ("EPS AR",        res["EPS"],       False),
        ("DME AR",        res["DME"],       False),
        ("Patient AR",    res["Patient"],   False),
    ]
    for ri, (label, vals, is_total) in enumerate(rows_data):
        excel_row = 3 + ri
        fill = _TOTAL_FILL if is_total else (_ALT_FILL if ri % 2 == 0 else _WHITE_FILL)
        lc = ws.cell(row=excel_row, column=1, value=label)
        lc.font = _F_BOLD; lc.fill = fill; lc.alignment = _LEFT; lc.border = _BORDER
        tar = total_ar(vals)
        tc2 = ws.cell(row=excel_row, column=2, value=round(tar, 2))
        tc2.font = _F_BOLD; tc2.fill = fill; tc2.number_format = _USD
        tc2.alignment = _RIGHT; tc2.border = _BORDER
        for bi, bv in enumerate(vals):
            vc = ws.cell(row=excel_row, column=3 + bi, value=round(float(bv), 2))
            vc.font = _F_NORMAL; vc.fill = fill; vc.number_format = _USD
            vc.alignment = _RIGHT; vc.border = _BORDER
        pct = pct_over_120(vals)
        pc = ws.cell(row=excel_row, column=9, value=round(pct, 4))
        pc.font = _F_BOLD if is_total else _F_NORMAL
        pc.fill = fill; pc.number_format = _PCT; pc.alignment = _CENTER; pc.border = _BORDER
        ws.row_dimensions[excel_row].height = 18
    ws.freeze_panes = "B3"
    return ws

# ── Core AGE_24 processor ─────────────────────────────────────────────────────
def process_bytes(file_bytes, start_date=None, end_date=None, report_date=None, orig_filename=""):
    buf = io.BytesIO(file_bytes)

    # Get sheet names (read-only, fast)
    try:
        wb_check = load_workbook(buf, read_only=True, data_only=True)
        all_sheets = wb_check.sheetnames
        wb_check.close()
    except Exception as e:
        return False, f"Cannot open file: {e}", None, None

    if not all_sheets:
        return False, "No sheets found in the workbook.", None, None

    raw_sheet_name = all_sheets[0]

    # Determine the report date
    if report_date:
        try:
            sheet_date = datetime.datetime.strptime(report_date, "%Y-%m-%d").date()
        except ValueError:
            return False, f"Invalid report date '{report_date}'.", None, None
    else:
        parsed = False
        for fmt in ("%m-%d-%Y", "%m/%d/%Y", "%Y-%m-%d"):
            try:
                sheet_date = datetime.datetime.strptime(raw_sheet_name.strip(), fmt).date()
                parsed = True
                break
            except ValueError:
                continue
        if not parsed:
            sheet_date = datetime.date.today()

    # Read raw data — dtype=str for fast initial load, then convert numerics in bulk
    try:
        buf.seek(0)
        df = pd.read_excel(
            buf,
            sheet_name=raw_sheet_name,
            header=0,
            engine="openpyxl",
            dtype=str,
        )
    except Exception as e:
        return False, f"Cannot read sheet '{raw_sheet_name}': {e}", None, None

    required_cols = AMOUNT_COLS + ["Financial_Class", "textbox18"]
    missing = [c for c in required_cols if c not in df.columns]
    if missing:
        return False, f"Missing columns: {missing}", None, None

    # Convert numeric columns once, up-front
    df["textbox18"] = pd.to_numeric(df["textbox18"], errors="coerce")
    for col in AMOUNT_COLS:
        df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0.0)

    # Filter: keep only textbox18 > 0
    df = df[df["textbox18"] > 0].copy()

    # Date range filter on Svc_Date
    date_filter_label = None
    if start_date or end_date:
        if "Svc_Date" not in df.columns:
            return False, "Column 'Svc_Date' not found. Date filtering requires it.", None, None
        df["Svc_Date"] = pd.to_datetime(df["Svc_Date"], errors="coerce")
        before = len(df)
        if start_date:
            df = df[df["Svc_Date"] >= pd.to_datetime(start_date)]
        if end_date:
            df = df[df["Svc_Date"] <= pd.to_datetime(end_date)]
        after = len(df)
        date_filter_label = f"{start_date or '—'}  →  {end_date or '—'}  ({after} of {before} rows)"

    filtered_count = len(df)
    fc = df["Financial_Class"]

    masks = {
        "Insurance": ~fc.isin(EXCLUDE_ALL),
        "Patient":    fc == "1-Self Pay",
        "WC":         fc.isin(WC_MVA_CATS),
        "EPS":        fc == "9-EPS",
        "DME":        fc == "11-DMERC",
    }
    res     = {k: bucket_sums(df, v) for k, v in masks.items()}
    overall = [sum(res[k][i] for k in res) for i in range(6)]

    # Free the dataframe — no longer needed
    del df

    # Build output workbook — load Summary from template
    try:
        wb_tmpl = load_summary_from_template()
    except (FileNotFoundError, ValueError) as tmpl_err:
        return False, f"Cannot load Summary template: {tmpl_err}", None, None

    wb = Workbook()
    wb.remove(wb.active)

    # Copy Summary sheet (styled)
    ws_src = wb_tmpl[SUMMARY_SHEET]
    ws_dst = wb.create_sheet(SUMMARY_SHEET, 0)
    for row in ws_src.iter_rows():
        for cell in row:
            nc = ws_dst.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                nc.font          = copy_cell_style(cell, "font")
                nc.border        = copy_cell_style(cell, "border")
                nc.fill          = copy_cell_style(cell, "fill")
                nc.number_format = cell.number_format
                nc.alignment     = copy_cell_style(cell, "alignment")
    for merge in ws_src.merged_cells.ranges:
        ws_dst.merge_cells(str(merge))
    for cl, cd in ws_src.column_dimensions.items():
        ws_dst.column_dimensions[cl].width = cd.width
    for ri, rd in ws_src.row_dimensions.items():
        ws_dst.row_dimensions[ri].height = rd.height
    wb_tmpl.close()

    # Copy raw sheet using read-only source — values only, no style (fast for large files)
    buf.seek(0)
    wb_orig = load_workbook(buf, read_only=True, data_only=True)
    ws_raw_src = wb_orig[raw_sheet_name]
    ws_raw_dst = wb.create_sheet(raw_sheet_name)
    for row in ws_raw_src.iter_rows():
        for cell in row:
            # read-only mode can yield EmptyCell objects with no row/column attrs
            try:
                r, c, v = cell.row, cell.column, cell.value
            except AttributeError:
                continue
            if v is not None:
                ws_raw_dst.cell(row=r, column=c, value=v)
    wb_orig.close()

    # Write calculated values into Summary sheet
    ws = wb[SUMMARY_SHEET]
    date_col = find_date_column(ws, sheet_date)
    if date_col is None:
        return False, f"Date {sheet_date} not found in Summary sheet header row.", None, None

    col_letter = get_column_letter(date_col)
    updates = {
        3:  total_ar(overall),          4:  overall[0],              5:  overall[1],
        6:  overall[2],                 7:  overall[3],              8:  overall[4],
        9:  overall[5],                 10: pct_over_120(overall),
        11: total_ar(res["Insurance"]), 12: res["Insurance"][0],     13: res["Insurance"][1],
        14: res["Insurance"][2],        15: res["Insurance"][3],     16: res["Insurance"][4],
        17: res["Insurance"][5],        18: pct_over_120(res["Insurance"]),
        19: total_ar(res["Patient"]),   20: res["Patient"][0],       21: res["Patient"][1],
        22: res["Patient"][2],          23: res["Patient"][3],       24: res["Patient"][4],
        25: res["Patient"][5],          26: pct_over_120(res["Patient"]),
        27: total_ar(res["WC"]),        28: res["WC"][0],            29: res["WC"][1],
        30: res["WC"][2],               31: res["WC"][3],            32: res["WC"][4],
        33: res["WC"][5],               34: pct_over_120(res["WC"]),
        35: total_ar(res["EPS"]),       36: res["EPS"][0],           37: res["EPS"][1],
        38: res["EPS"][2],              39: res["EPS"][3],           40: res["EPS"][4],
        41: res["EPS"][5],              42: pct_over_120(res["EPS"]),
        43: total_ar(res["DME"]),       44: res["DME"][0],           45: res["DME"][1],
        46: res["DME"][2],              47: res["DME"][3],           48: res["DME"][4],
        49: res["DME"][5],              50: pct_over_120(res["DME"]),
    }
    for row, val in updates.items():
        cell = ws.cell(row=row, column=date_col)
        cell.value         = round(float(val), 4)
        cell.number_format = "0.00%" if row in PCT_ROWS else "#,##0.00"

    write_report_sheet(wb, overall, res, start_date=start_date, end_date=end_date)

    out_buf   = io.BytesIO()
    wb.save(out_buf)
    out_bytes = out_buf.getvalue()
    del out_buf

    summary = {
        "sheet":       raw_sheet_name,
        "date":        str(sheet_date),
        "report_date": str(sheet_date),
        "col_letter":  col_letter,
        "rows":        filtered_count,
        "date_filter": date_filter_label,
        "categories": [
            {"name": "Overall Aging", "total": round(total_ar(overall),          2), "pct": round(pct_over_120(overall)*100,          2)},
            {"name": "Insurance AR",  "total": round(total_ar(res["Insurance"]), 2), "pct": round(pct_over_120(res["Insurance"])*100, 2)},
            {"name": "Patient AR",    "total": round(total_ar(res["Patient"]),   2), "pct": round(pct_over_120(res["Patient"])*100,   2)},
            {"name": "WC & MVA",      "total": round(total_ar(res["WC"]),        2), "pct": round(pct_over_120(res["WC"])*100,        2)},
            {"name": "EPS AR",        "total": round(total_ar(res["EPS"]),       2), "pct": round(pct_over_120(res["EPS"])*100,       2)},
            {"name": "DME AR",        "total": round(total_ar(res["DME"]),       2), "pct": round(pct_over_120(res["DME"])*100,       2)},
        ]
    }
    return True, "Success", out_bytes, summary

# ── Global error handlers ─────────────────────────────────────────────────────
@app.errorhandler(404)
def not_found(e):
    return jsonify({"success": False, "message": "Endpoint not found."}), 404

@app.errorhandler(500)
def internal_error(e):
    return jsonify({"success": False, "message": f"Internal server error: {str(e)}"}), 500

@app.errorhandler(413)
def too_large(e):
    return jsonify({"success": False, "message": "File too large. Maximum upload size is 200 MB."}), 413

# ── Routes ────────────────────────────────────────────────────────────────────
@app.route("/")
def index():
    return render_template("index.html")

@app.route("/process", methods=["POST"])
def process_route():
    try:
        if "file" not in request.files:
            return jsonify({"success": False, "message": "No file uploaded."}), 400
        f = request.files["file"]
        if not f.filename.endswith(".xlsx"):
            return jsonify({"success": False, "message": "Please upload an .xlsx file."}), 400
        fname_lower = f.filename.lower()
        if "age_24" not in fname_lower and "age24" not in fname_lower:
            return jsonify({"success": False, "message": "File name must contain 'Age_24'."}), 400

        start_date  = request.form.get("start_date")  or None
        end_date    = request.form.get("end_date")    or None
        report_date = request.form.get("report_date") or None

        file_bytes = f.read()
        success, message, out_bytes, summary = process_bytes(
            file_bytes, start_date, end_date,
            report_date=report_date, orig_filename=f.filename
        )
        if not success:
            return jsonify({"success": False, "message": message}), 400

        token = _save_output(out_bytes, f.filename)
        return jsonify({"success": True, "token": token, "summary": summary})

    except Exception as e:
        traceback.print_exc()
        return jsonify({"success": False, "message": f"Server error: {str(e)}"}), 500

@app.route("/download/<token>")
def download(token):
    entry = _get_output(token)
    if not entry:
        return jsonify({"success": False, "message": "File not found or expired."}), 404
    tmp_path, orig_name = entry
    return send_file(
        tmp_path,
        as_attachment=True,
        download_name=f"{Path(orig_name).stem}_updated.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

# ═══════════════════════════════════════════════════════════════════════════════
# ── PAY_41 Payment Summary Generator ──────────────────────────────────────────
# ═══════════════════════════════════════════════════════════════════════════════

MONTH_FMT    = "%b-%y"
_HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
_SUBHDR_FILL = PatternFill("solid", fgColor="2E75B6")
_PAY_TOTAL_FILL = PatternFill("solid", fgColor="BDD7EE")
_PAY_ALT_FILL   = PatternFill("solid", fgColor="DEEAF1")
_PAY_WHITE_FILL = PatternFill("solid", fgColor="FFFFFF")
_WHITE_BOLD  = Font(name="Arial", bold=True, color="FFFFFF", size=10)
_NORMAL      = Font(name="Arial", size=10)
_NORMAL_BOLD = Font(name="Arial", bold=True, size=10)
_THIN_PAY    = Side(style="thin", color="B0B0B0")
_BORDER_THIN = Border(left=_THIN_PAY, right=_THIN_PAY, top=_THIN_PAY, bottom=_THIN_PAY)
_CENTER_PAY  = Alignment(horizontal="center", vertical="center", wrap_text=True)
_RIGHT_PAY   = Alignment(horizontal="right",  vertical="center")
_LEFT_PAY    = Alignment(horizontal="left",   vertical="center")
_USD_FMT     = '"$"#,##0.00_);("$"#,##0.00)'
_YEAR_COLOURS = ["4472C4","ED7D31","A9D18E","FF0000","FFC000","70AD47","5B9BD5","264478"]

HCPCS_DME_CODES = {
    "A9150","A7003","A6449","A4450","A6250","A6450","A4615","A6454","A4467","A6448",
    "A6216","A4314","A4452","A6222","A4566","A4550","E0114","E0110",
    "L3809","L1833","L4361","L3908","L1830","Q4049","L3660","L0642","L4350",
    "L1902","L4386","L3260","L0641","A4565","Q4018","L0120","L1820","L4360",
    "L3925","Q4020","L0174","Q4024","Q4012","L3670","Q4006","Q4042","L3980",
    "L3807","Q4022","Q4046","L3984","L1810","L3923","L0172",
}

def assign_pay_category_vectorised(df):
    """Vectorised — avoids slow row-by-row apply()."""
    fc  = df["Financial_Class"].fillna("").astype(str)
    pc  = df["Proc_Code"].fillna("").astype(str).str.strip()
    cat = pd.Series(np.nan, index=df.index, dtype="object")

    cat = cat.where(fc != "11-DMERC", "DME Medicare")
    dme_comm_mask = cat.isna() & pc.isin(HCPCS_DME_CODES)
    cat = cat.where(~dme_comm_mask, "DME Commercial")
    wc_mask  = cat.isna() & fc.str.contains(r"WC|MVA", case=False, regex=True, na=False)
    cat = cat.where(~wc_mask, "WC & MVA")
    eps_mask = cat.isna() & fc.str.contains(r"EMP", case=False, regex=True, na=False)
    cat = cat.where(~eps_mask, "EPS")
    return cat

def build_pay_pivot(df_cat):
    df = df_cat[["Svc_Date", "Trans_Date", "Payment"]].copy()
    df["dos_month"]   = df["Svc_Date"].dt.to_period("M")
    df["trans_month"] = df["Trans_Date"].dt.to_period("M")
    pivot = df.groupby(["dos_month","trans_month"])["Payment"].sum().unstack("trans_month")
    pivot = pivot.sort_index()
    pivot.columns = pivot.columns.sort_values()
    return pivot

def _write_pay_sheet(wb, category, pivot):
    safe_name = re.sub(r'[\\/*?:\[\]]', "_", category)[:31]
    ws = wb.create_sheet(title=safe_name)
    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=len(pivot.columns)+2)
    tc = ws.cell(row=1, column=1, value=f"{category} Transaction Summary")
    tc.font = Font(name="Arial", bold=True, color="FFFFFF", size=13)
    tc.fill = _HEADER_FILL; tc.alignment = _CENTER_PAY

    header_row   = 3
    trans_months = list(pivot.columns)
    col_labels   = [pd.Timestamp(str(p)).strftime(MONTH_FMT) for p in trans_months]
    total_col    = len(col_labels) + 2

    h = ws.cell(row=header_row, column=1, value="DOS")
    h.font = _WHITE_BOLD; h.fill = _SUBHDR_FILL; h.alignment = _CENTER_PAY
    ws.column_dimensions["A"].width = 12

    for ci, lbl in enumerate(col_labels, start=2):
        c = ws.cell(row=header_row, column=ci, value=lbl)
        c.font = _WHITE_BOLD; c.fill = _SUBHDR_FILL; c.alignment = _CENTER_PAY
        ws.column_dimensions[get_column_letter(ci)].width = 11

    tc2 = ws.cell(row=header_row, column=total_col, value="Total")
    tc2.font = _WHITE_BOLD; tc2.fill = _SUBHDR_FILL; tc2.alignment = _CENTER_PAY
    ws.column_dimensions[get_column_letter(total_col)].width = 12

    data_start_row = header_row + 1
    for ri, (dos_period, row_data) in enumerate(pivot.iterrows()):
        excel_row = data_start_row + ri
        fill      = _PAY_ALT_FILL if ri % 2 == 0 else _PAY_WHITE_FILL
        dos_lbl   = pd.Timestamp(str(dos_period)).strftime(MONTH_FMT)
        dc = ws.cell(row=excel_row, column=1, value=dos_lbl)
        dc.font = _NORMAL_BOLD; dc.fill = fill; dc.alignment = _LEFT_PAY; dc.border = _BORDER_THIN
        for ci, period in enumerate(trans_months, start=2):
            val = row_data.get(period, float("nan"))
            vc  = ws.cell(row=excel_row, column=ci, value=None if pd.isna(val) else val)
            vc.fill = fill; vc.alignment = _RIGHT_PAY
            vc.number_format = _USD_FMT; vc.border = _BORDER_THIN; vc.font = _NORMAL
        fv  = get_column_letter(2); lv = get_column_letter(total_col - 1)
        tc3 = ws.cell(row=excel_row, column=total_col,
                      value=f"=SUM({fv}{excel_row}:{lv}{excel_row})")
        tc3.fill = _PAY_TOTAL_FILL; tc3.alignment = _RIGHT_PAY
        tc3.number_format = _USD_FMT; tc3.border = _BORDER_THIN; tc3.font = _NORMAL_BOLD

    total_row = data_start_row + len(pivot)
    gt = ws.cell(row=total_row, column=1, value="Grand Total")
    gt.font = _WHITE_BOLD; gt.fill = _HEADER_FILL
    gt.alignment = _LEFT_PAY; gt.border = _BORDER_THIN
    for ci in range(2, total_col + 1):
        cl = get_column_letter(ci)
        ct = ws.cell(row=total_row, column=ci,
                     value=f"=SUM({cl}{data_start_row}:{cl}{total_row-1})")
        ct.font = _WHITE_BOLD; ct.fill = _HEADER_FILL
        ct.alignment = _RIGHT_PAY; ct.number_format = _USD_FMT; ct.border = _BORDER_THIN

    chart_top_row    = total_row + 2
    chart_col_idx    = 1
    chart_width_cols = (total_col + 1) // 2 + 1
    for s_idx, col_idx in enumerate(range(2, total_col)):
        trans_label = ws.cell(row=header_row, column=col_idx).value
        if not trans_label:
            continue
        chart = LineChart()
        chart.title  = trans_label; chart.style = 10
        chart.height = 12;          chart.width = 22
        chart.y_axis.numFmt = '"$"#,##0.00;("$"#,##0.00)'
        chart.y_axis.majorGridlines = None
        chart.x_axis.tickLblPos = "low"
        data_ref = Reference(ws, min_col=col_idx, max_col=col_idx,
                             min_row=data_start_row, max_row=total_row-1)
        ser = Series(data_ref, title=trans_label)
        colour = _YEAR_COLOURS[s_idx % len(_YEAR_COLOURS)]
        ser.graphicalProperties.line.solidFill = colour
        ser.graphicalProperties.line.width     = 22000
        ser.marker.symbol = "circle"; ser.marker.size = 5
        ser.marker.graphicalProperties.solidFill      = colour
        ser.marker.graphicalProperties.line.solidFill = colour
        from openpyxl.chart.label import DataLabelList
        ser.dLbls = DataLabelList()
        ser.dLbls.showVal = True; ser.dLbls.showLegendKey = False
        ser.dLbls.showCatName = False; ser.dLbls.showSerName = False
        ser.dLbls.numFmt = '"$"#,##0.00;("$"#,##0.00)'; ser.dLbls.position = "t"
        chart.append(ser)
        chart.set_categories(Reference(ws, min_col=1,
                                       min_row=data_start_row, max_row=total_row-1))
        ws.add_chart(chart, f"{get_column_letter(chart_col_idx)}{chart_top_row}")
        if chart_col_idx == 1:
            chart_col_idx = chart_width_cols
        else:
            chart_col_idx = 1; chart_top_row += 28
    ws.freeze_panes = "B4"

def process_pay_summary(file_bytes):
    buf = io.BytesIO(file_bytes)
    try:
        tmp_wb = load_workbook(buf, read_only=True, data_only=True)
        pay_sheet_name = tmp_wb.sheetnames[0]
        tmp_wb.close()
    except Exception as e:
        return False, f"Cannot open file: {e}", None, None

    buf.seek(0)
    try:
        df = pd.read_excel(
            buf,
            sheet_name=pay_sheet_name,
            engine="openpyxl",
            usecols=["Svc_Date", "Trans_Date", "Proc_Code", "Financial_Class", "Payment"],
            parse_dates=["Svc_Date", "Trans_Date"],
        )
    except Exception as e:
        return False, f"Cannot read sheet '{pay_sheet_name}': {e}", None, None

    df["Payment"] = pd.to_numeric(df["Payment"], errors="coerce").fillna(0.0)
    df["Category"] = assign_pay_category_vectorised(df)

    TARGET     = ["DME Medicare", "DME Commercial", "WC & MVA", "EPS"]
    df         = df[df["Category"].isin(TARGET)].copy()
    categories = [c for c in TARGET if c in df["Category"].values]

    wb = Workbook(); wb.remove(wb.active)
    summary_cats = []
    for cat in categories:
        df_cat = df[df["Category"] == cat]
        pivot  = build_pay_pivot(df_cat)
        _write_pay_sheet(wb, cat, pivot)
        summary_cats.append({"name": cat, "dos_rows": len(pivot), "trans_cols": len(pivot.columns)})

    out_buf   = io.BytesIO()
    wb.save(out_buf)
    out_bytes = out_buf.getvalue()
    del out_buf, df

    return True, "Success", out_bytes, {
        "categories":  summary_cats,
        "hcpcs_count": len(HCPCS_DME_CODES),
        "pay_sheet":   pay_sheet_name,
    }

@app.route("/summary")
def summary_page():
    return render_template("summary.html")

@app.route("/process-summary", methods=["POST"])
def process_summary_route():
    try:
        if "file" not in request.files:
            return jsonify({"success": False, "message": "No file uploaded."}), 400
        f = request.files["file"]
        if not f.filename.endswith(".xlsx"):
            return jsonify({"success": False, "message": "Please upload an .xlsx file."}), 400
        if "pay_41" not in f.filename.lower() and "pay41" not in f.filename.lower():
            return jsonify({"success": False, "message": "File name must contain 'PAY_41'."}), 400

        file_bytes = f.read()
        success, message, out_bytes, summary = process_pay_summary(file_bytes)
        if not success:
            return jsonify({"success": False, "message": message}), 400

        token = _save_output(out_bytes, f.filename)
        return jsonify({"success": True, "token": token, "summary": summary})

    except Exception as e:
        traceback.print_exc()
        return jsonify({"success": False, "message": f"Server error: {str(e)}"}), 500

@app.route("/download-summary/<token>")
def download_summary(token):
    entry = _get_output(token)
    if not entry:
        return jsonify({"success": False, "message": "File not found or expired."}), 404
    tmp_path, orig_name = entry
    return send_file(
        tmp_path,
        as_attachment=True,
        download_name=f"{Path(orig_name).stem}_PaymentSummary.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

# ═══════════════════════════════════════════════════════════════════════════════
if __name__ == "__main__":
    import socket
    hostname = socket.gethostname()
    try:
        local_ip = socket.gethostbyname(hostname)
    except Exception:
        local_ip = "your-pc-ip"
    print(f"\n{'='*55}")
    print(f"  AGE_24 / PAY_41 Web Processor is running!")
    print(f"  Open in browser on this PC : http://localhost:5000")
    print(f"  Open from other PCs on WiFi: http://{local_ip}:5000")
    print(f"{'='*55}\n")
    app.run(host="0.0.0.0", port=5000, debug=False, threaded=True)
